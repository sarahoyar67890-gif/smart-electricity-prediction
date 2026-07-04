"""
Report Generation Module
Generates: Dataset Summary, Model Performance Report, Energy Consumption
Report, Monthly Usage Forecast, Recommendation Summary.
Exports as: PDF (reportlab), Excel (openpyxl), and CSV.
"""

import pandas as pd
import json
import os
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer, Table,
                                 TableStyle, Image)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA_PATH = os.path.join(PROJECT_ROOT, "data", "electricity_dataset_clean.csv")
REPORTS_DIR = os.path.join(PROJECT_ROOT, "reports")
EDA_DIR = f"{REPORTS_DIR}/eda"
MODEL_REPORT_JSON = f"{REPORTS_DIR}/model_performance_report.json"

PINK_HEX = "#F8B8D0"       # for reportlab (needs #)
DARK_PINK_HEX = "#E75480"  # for reportlab (needs #)
PINK_HEX_XL = "FFF8B8D0"       # for openpyxl (ARGB, no #)
DARK_PINK_HEX_XL = "FFE75480"  # for openpyxl (ARGB, no #)


def generate_pdf_report():
    df = pd.read_csv(DATA_PATH)
    with open(MODEL_REPORT_JSON) as f:
        model_report = json.load(f)

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("TitlePink", parent=styles["Title"],
                                  textColor=colors.HexColor(DARK_PINK_HEX), fontSize=22)
    heading_style = ParagraphStyle("HeadingPink", parent=styles["Heading2"],
                                    textColor=colors.HexColor(DARK_PINK_HEX))
    normal = styles["Normal"]

    doc = SimpleDocTemplate(f"{REPORTS_DIR}/Project_Report.pdf", pagesize=A4,
                             topMargin=1.5*cm, bottomMargin=1.5*cm)
    story = []

    story.append(Paragraph("⚡ Smart Electricity Consumption Prediction", title_style))
    story.append(Paragraph("Energy Optimization System — Project Report", styles["Heading3"]))
    story.append(Spacer(1, 0.5*cm))

    # Dataset Summary
    story.append(Paragraph("1. Dataset Summary", heading_style))
    ds_text = f"""
    Original synthetic dataset created for this project (no public dataset used).<br/>
    Total Records: {df.shape[0]} &nbsp;&nbsp; | &nbsp;&nbsp; Total Features: {df.shape[1]-1}<br/>
    Target Variable: Daily_Electricity_Consumption_kWh<br/>
    Mean Consumption: {df['Daily_Electricity_Consumption_kWh'].mean():.2f} kWh/day<br/>
    Std Deviation: {df['Daily_Electricity_Consumption_kWh'].std():.2f} kWh
    """
    story.append(Paragraph(ds_text, normal))
    story.append(Spacer(1, 0.4*cm))

    # Model Performance
    story.append(Paragraph("2. Model Performance Report", heading_style))
    story.append(Paragraph(f"Best Performing Model: <b>{model_report['best_model']}</b>", normal))
    story.append(Spacer(1, 0.2*cm))

    table_data = [["Model", "MAE", "MSE", "RMSE", "R2 Score"]]
    for name, m in model_report["metrics"].items():
        table_data.append([name, str(m["MAE"]), str(m["MSE"]), str(m["RMSE"]), str(m["R2_Score"])])

    t = Table(table_data, colWidths=[4.5*cm, 2.5*cm, 2.5*cm, 2.5*cm, 2.5*cm])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(DARK_PINK_HEX)),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor(PINK_HEX)),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.white),
        ("ALIGN", (1, 0), (-1, -1), "CENTER"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
    ]))
    story.append(t)
    story.append(Spacer(1, 0.5*cm))

    # EDA visuals
    story.append(Paragraph("3. Exploratory Data Analysis", heading_style))
    for img_name, caption in [
        ("correlation_heatmap.png", "Correlation Heatmap"),
        ("model_comparison.png", "Model Comparison (R2 / RMSE)"),
        ("actual_vs_predicted.png", "Actual vs Predicted Consumption"),
        ("feature_importance.png", "Feature Importance"),
    ]:
        img_path = f"{EDA_DIR}/{img_name}"
        if os.path.exists(img_path):
            story.append(Paragraph(caption, styles["Heading4"]))
            story.append(Image(img_path, width=14*cm, height=9*cm))
            story.append(Spacer(1, 0.3*cm))

    # Business Insights
    story.append(Paragraph("4. Energy Consumption Insights", heading_style))
    insights_path = f"{EDA_DIR}/business_insights.txt"
    if os.path.exists(insights_path):
        with open(insights_path) as f:
            insight_text = f.read().replace("\n", "<br/>")
        story.append(Paragraph(insight_text, normal))
    story.append(Spacer(1, 0.4*cm))

    # Recommendation Summary
    story.append(Paragraph("5. Energy Optimization Recommendations", heading_style))
    recs = """
    • Shift heavy appliance usage (AC, washing machine, iron) to off-peak hours (before 5 PM or after 10 PM).<br/>
    • Households using AC more than 6 hrs/day can save 15-25% by raising thermostat by 2°C.<br/>
    • Bungalows and independent houses show the highest consumption — insulation upgrades recommended.<br/>
    • Summer/Monsoon months require proactive energy budgeting due to cooling load spikes.<br/>
    • Standby/base load (chargers, routers) contributes a steady baseline — smart plugs recommended.
    """
    story.append(Paragraph(recs, normal))
    story.append(Spacer(1, 0.4*cm))

    # Challenges & Future Work
    story.append(Paragraph("6. Challenges Faced & Future Improvements", heading_style))
    challenges = """
    <b>Challenges:</b> Designing a realistic synthetic dataset that reflects genuine appliance-usage
    correlations (e.g., AC hours vs temperature) without access to real smart-meter data required
    careful domain-driven simulation rather than random generation.<br/><br/>
    <b>Future Improvements:</b> Integrate a live Weather API for real-time temperature-based prediction,
    deploy via Docker for production use, add real smart-meter data collection, and extend the model
    to multi-day/weekly forecasting.
    """
    story.append(Paragraph(challenges, normal))

    doc.build(story)
    print(f"PDF report saved: {REPORTS_DIR}/Project_Report.pdf")


def generate_excel_report():
    df = pd.read_csv(DATA_PATH)
    with open(MODEL_REPORT_JSON) as f:
        model_report = json.load(f)

    wb = Workbook()
    header_fill = PatternFill(start_color=DARK_PINK_HEX_XL, end_color=DARK_PINK_HEX_XL, fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    row_fill = PatternFill(start_color=PINK_HEX_XL, end_color=PINK_HEX_XL, fill_type="solid")

    # Sheet 1: Dataset Summary
    ws1 = wb.active
    ws1.title = "Dataset Summary"
    ws1.append(["Metric", "Value"])
    summary_rows = [
        ("Total Records", df.shape[0]),
        ("Total Features", df.shape[1] - 1),
        ("Mean Consumption (kWh)", round(df["Daily_Electricity_Consumption_kWh"].mean(), 2)),
        ("Median Consumption (kWh)", round(df["Daily_Electricity_Consumption_kWh"].median(), 2)),
        ("Std Dev (kWh)", round(df["Daily_Electricity_Consumption_kWh"].std(), 2)),
        ("Min Consumption (kWh)", round(df["Daily_Electricity_Consumption_kWh"].min(), 2)),
        ("Max Consumption (kWh)", round(df["Daily_Electricity_Consumption_kWh"].max(), 2)),
    ]
    for row in summary_rows:
        ws1.append(row)
    for cell in ws1[1]:
        cell.fill = header_fill
        cell.font = header_font
    ws1.column_dimensions["A"].width = 28
    ws1.column_dimensions["B"].width = 18

    # Sheet 2: Model Performance
    ws2 = wb.create_sheet("Model Performance")
    ws2.append(["Model", "MAE", "MSE", "RMSE", "R2 Score"])
    for name, m in model_report["metrics"].items():
        ws2.append([name, m["MAE"], m["MSE"], m["RMSE"], m["R2_Score"]])
    for cell in ws2[1]:
        cell.fill = header_fill
        cell.font = header_font
    for col in "ABCDE":
        ws2.column_dimensions[col].width = 18

    # Sheet 3: Monthly Forecast (based on average daily * 30, per house type)
    ws3 = wb.create_sheet("Monthly Forecast")
    ws3.append(["House Type", "Avg Daily (kWh)", "Est. Monthly (kWh)", "Est. Monthly Bill (PKR)"])
    RATE_PKR_PER_KWH = 45  # approximate average slab rate for estimation
    house_avg = df.groupby("House_Type")["Daily_Electricity_Consumption_kWh"].mean()
    for house, avg in house_avg.items():
        monthly = avg * 30
        bill = monthly * RATE_PKR_PER_KWH
        ws3.append([house, round(avg, 2), round(monthly, 2), round(bill, 0)])
    for cell in ws3[1]:
        cell.fill = header_fill
        cell.font = header_font
    for col in "ABCD":
        ws3.column_dimensions[col].width = 22

    # Sheet 4: Full cleaned dataset (first 200 rows sample for size)
    ws4 = wb.create_sheet("Dataset Sample")
    ws4.append(list(df.columns))
    for _, row in df.head(200).iterrows():
        ws4.append(list(row))
    for cell in ws4[1]:
        cell.fill = header_fill
        cell.font = header_font

    wb.save(f"{REPORTS_DIR}/Energy_Report.xlsx")
    print(f"Excel report saved: {REPORTS_DIR}/Energy_Report.xlsx")


def generate_csv_reports():
    df = pd.read_csv(DATA_PATH)
    # Consumption report by season + house type
    pivot = df.pivot_table(index="Season", columns="House_Type",
                            values="Daily_Electricity_Consumption_kWh", aggfunc="mean").round(2)
    pivot.to_csv(f"{REPORTS_DIR}/energy_consumption_report.csv")
    print(f"CSV report saved: {REPORTS_DIR}/energy_consumption_report.csv")


if __name__ == "__main__":
    generate_pdf_report()
    generate_excel_report()
    generate_csv_reports()
